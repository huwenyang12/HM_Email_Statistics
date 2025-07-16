# excel_exporter.py
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import pandas as pd

def get_excel_styles():
    return {
        'red_fill': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),           # 浅红
        'green_fill': PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid"),         # 浅绿
        'yellow_fill': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),        # 浅黄
        'gray_fill': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),          # 浅灰
        'header_fill': PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),        # 表头灰
        'center_align': Alignment(horizontal="center", vertical="center"),
        'left_align': Alignment(horizontal="left", vertical="center"),
        'thin_border': Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    }

def set_column_widths(worksheet, headers):
    worksheet.column_dimensions['A'].width = 8
    worksheet.column_dimensions['B'].width = 30
    for i in range(2, len(headers)):
        col_letter = chr(ord('A') + i)
        header = headers[i]
        if '巡检时间' in header:
            worksheet.column_dimensions[col_letter].width = 12
        elif '巡检结果' in header:
            worksheet.column_dimensions[col_letter].width = 15
        elif header == '成功率':
            worksheet.column_dimensions[col_letter].width = 10
        elif header == '备注':
            worksheet.column_dimensions[col_letter].width = 40  

def apply_excel_styles(worksheet, headers):
    styles = get_excel_styles()

    # 设置表头样式
    for cell in worksheet[1]:
        cell.fill = styles['header_fill']
        cell.alignment = styles['center_align']
        cell.border = styles['thin_border']

    time_col_indices = [idx for idx, h in enumerate(headers) if h.startswith("巡检时间")]

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for col_idx, cell in enumerate(row):
            cell.border = styles['thin_border']
            cell.alignment = styles['left_align'] if col_idx == 1 else styles['center_align']

            if col_idx in time_col_indices:
                cell.fill = PatternFill(fill_type=None)
                continue

            if isinstance(cell.value, str):
                val = cell.value.strip()
                if val == "巡检失败":
                    cell.fill = styles['red_fill']
                elif val == "巡检中":
                    cell.fill = styles['yellow_fill']
                elif val == "暂未巡检":
                    cell.fill = styles['gray_fill']
                elif ":" in val:  # 成功时间点
                    cell.fill = styles['green_fill']

def append_summary_success_rate(worksheet, headers):
    try:
        success_col_idx = headers.index("成功率") + 1
    except ValueError:
        print("未找到成功率列")
        return

    total_success = total_fail = 0
    result_col_indices = [idx for idx, h in enumerate(headers) if "巡检结果" in h]

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for idx in result_col_indices:
            cell_value = row[idx].value
            if isinstance(cell_value, str):
                value = cell_value.strip()
                if value == "巡检失败":
                    total_fail += 1
                elif ":" in value:
                    total_success += 1

    total_checked = total_success + total_fail
    success_rate = round(total_success / total_checked * 100, 2) if total_checked else 0.0
    success_rate_str = f"{success_rate}%"

    target_row = worksheet.max_row + 1
    cell = worksheet.cell(row=target_row, column=success_col_idx, value=success_rate_str)

    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def export_to_excel(df, headers, file_path):
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='巡检统计', index=False)
        worksheet = writer.sheets['巡检统计']
        set_column_widths(worksheet, headers)
        apply_excel_styles(worksheet, headers)
        append_summary_success_rate(worksheet, headers)
